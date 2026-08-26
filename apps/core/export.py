"""CSV + XLSX export helpers.

CSV: UTF-8 BOM + semicolon delimiter for Latvian Excel, with value formatting
and a spreadsheet formula-injection guard. P17 adds ``prepare_export_cell``
(shared cell formatting + guard) and ``xlsx_response`` (in-memory XLSX writer
via openpyxl).
"""

from __future__ import annotations

import csv
import datetime
import io
from collections.abc import Iterable, Sequence
from typing import Any

from django.http import HttpResponse

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover — openpyxl is a required dep
    Workbook = None  # type: ignore[assignment]

_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _format(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "jā" if value else "nē"
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, datetime.date):
        return value.isoformat()
    return str(value)


def _guard(text: str) -> str:
    if text and text[0] in _INJECTION_PREFIXES:
        return "'" + text
    return text


def prepare_export_cell(value: Any) -> str:
    """Shared cell formatter: render via ``_format``, then formula-guard.

    Used by both ``csv_response`` and ``xlsx_response``. Strings starting with
    ``=``, ``+``, ``-``, ``@``, tab, or carriage return are prefixed with
    ``'`` so a downstream Excel does not execute the cell as a formula.
    """
    return _guard(_format(value))


def csv_response(*, filename: str, columns: Sequence[str], rows: Iterable[Sequence[Any]]) -> HttpResponse:
    buf = io.StringIO()
    buf.write("﻿")  # UTF-8 BOM so Latvian Excel detects UTF-8
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    writer.writerow(list(columns))
    for row in rows:
        writer.writerow([prepare_export_cell(cell) for cell in row])
    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def xlsx_response(
    *, filename: str, columns: Sequence[str], rows: Iterable[Sequence[Any]]
) -> HttpResponse:
    """In-memory XLSX writer — bold headers, widths clamped 10–40."""
    if Workbook is None:  # pragma: no cover — dep guard
        raise RuntimeError("openpyxl is required for xlsx_response")
    wb = Workbook()
    ws = wb.active
    ws.append(list(columns))
    # Bold the header row (row 1).
    from openpyxl.styles import Font

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    for row in rows:
        ws.append([prepare_export_cell(cell) for cell in row])
    # Clamp column widths to [10, 40] based on header label length.
    for idx, header in enumerate(columns, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        width = max(10, min(40, len(str(header)) + 2))
        ws.column_dimensions[letter].width = float(width)
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
