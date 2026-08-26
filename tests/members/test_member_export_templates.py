"""P17 — configurable member export: column registry, template model, service.

Every test describes the desired API of code that does not exist yet
(``apps.members.exports`` registry additions, ``MemberExportTemplate`` model,
``apps.members.export_templates`` service, ``apps.members.forms`` widgets).
"""

import csv
import datetime
import io

import pytest
from django.core.exceptions import ValidationError
from django.db import connection
from django.test.utils import CaptureQueriesContext

from apps.members.models import Member

pytestmark = pytest.mark.django_db


def _make_agreement(member, *, state="generated", is_current=True):
    from django.utils import timezone

    from apps.agreements.models import Agreement

    return Agreement.objects.create(
        member=member,
        is_current=is_current,
        state=state,
        generated_at=timezone.now(),
    )


def _member_with_agreement(guardian, *, group=None, state="generated", full_name="Member"):
    from django.utils import timezone

    from apps.agreements.models import Agreement
    from apps.members.models import Member

    member = Member.objects.create(
        full_name=full_name, guardian=guardian, training_group=group
    )
    Agreement.objects.create(
        member=member, state=state, is_current=True, generated_at=timezone.now()
    )
    return member


# ---------------------------------------------------------------------------
# Column registry
# ---------------------------------------------------------------------------


def test_column_registry_has_all_keys():
    from apps.members.exports import COLUMN_REGISTRY

    expected_keys = [
        "member_full_name",
        "member_personal_id",
        "member_birth_date",
        "guardian_name",
        "guardian_email",
        "guardian_phone",
        "guardian_address",
        "agreement_state",
        "agreement_signed_at",
        "training_group_name",
    ]
    assert list(COLUMN_REGISTRY.keys()) == expected_keys


def test_column_registry_has_column_spec_structure():
    from apps.members.exports import COLUMN_REGISTRY, ColumnSpec

    for key, col in COLUMN_REGISTRY.items():
        assert isinstance(col, ColumnSpec)
        assert col.key == key
        assert isinstance(col.label, str) and len(col.label) > 0
        assert callable(col.reader)
        assert isinstance(col.sensitive, bool)


def test_column_registry_has_latvian_labels():
    from apps.members.exports import COLUMN_REGISTRY

    expected = {
        "member_full_name": "Biedra vārds, uzvārds",
        "member_personal_id": "Biedra personas kods",
        "member_birth_date": "Biedra dzimšanas datums",
        "guardian_name": "Vecāka vārds, uzvārds",
        "guardian_email": "Vecāka e-pasts",
        "guardian_phone": "Vecāka tālrunis",
        "guardian_address": "Vecāka adrese",
        "agreement_state": "Līguma statuss",
        "agreement_signed_at": "Līguma parakstīšanas datums",
        "training_group_name": "Treniņu grupa",
    }
    assert {k: COLUMN_REGISTRY[k].label for k in expected} == expected


def test_column_registry_sensitive_subset():
    from apps.members.exports import COLUMN_REGISTRY, SENSITIVE_KEYS

    expected_sensitive = {
        "member_personal_id",
        "guardian_email",
        "guardian_phone",
        "guardian_address",
    }
    assert SENSITIVE_KEYS == expected_sensitive
    for key, col in COLUMN_REGISTRY.items():
        assert col.sensitive == (key in expected_sensitive)


def test_column_registry_readers_are_pure():
    from apps.members.exports import COLUMN_REGISTRY

    member = Member(pk=1, full_name="Test")
    for key, col in COLUMN_REGISTRY.items():
        with CaptureQueriesContext(connection) as ctx:
            result = col.reader(member)
        assert len(ctx.captured_queries) == 0, f"reader {key} issued a query"
        assert isinstance(result, (str, type(None), int, float, datetime.date))


def test_readers_emit_em_dash_for_missing_group_and_agreement():
    from apps.members.exports import COLUMN_REGISTRY

    # Bare unsaved Member: no training group and no prefetched current
    # agreements. The readers must render the "—" placeholder without
    # issuing any ORM query; agreement_signed_at stays None.
    member = Member(full_name="Bez grupas")

    with CaptureQueriesContext(connection) as ctx:
        assert COLUMN_REGISTRY["agreement_state"].reader(member) == "—"
        assert COLUMN_REGISTRY["training_group_name"].reader(member) == "—"
        assert COLUMN_REGISTRY["agreement_signed_at"].reader(member) is None
    assert len(ctx.captured_queries) == 0


def test_column_registry_readers_return_real_values(guardian, training_group_a):
    import datetime

    from django.utils import timezone

    from apps.agreements.models import Agreement
    from apps.members.exports import COLUMN_REGISTRY
    from apps.members.models import Member

    member = Member.objects.create(
        full_name="Jānis Bērziņš",
        personal_id="151210-22222",
        birth_date="2015-12-10",
        guardian=guardian,
        training_group=training_group_a,
    )
    signed_at = timezone.now()
    member._current_export_agreements = [
        Agreement(
            member=member,
            state="signed",
            generated_at=timezone.now(),
            signed_at=signed_at,
        )
    ]

    expected = {
        "member_full_name": "Jānis Bērziņš",
        "member_personal_id": "151210-22222",
        "member_birth_date": datetime.date(2015, 12, 10),
        "guardian_name": "Anna Bērziņa",
        "guardian_email": "anna@example.test",
        "guardian_phone": "+37120000000",
        "guardian_address": "Rīgas iela 1, Cēsis",
        "agreement_state": "Parakstīts",
        "agreement_signed_at": signed_at,
        "training_group_name": "U10 A",
    }

    with CaptureQueriesContext(connection) as ctx:
        for key, col in COLUMN_REGISTRY.items():
            assert col.reader(member) == expected[key], key
    assert len(ctx.captured_queries) == 0


# ---------------------------------------------------------------------------
# MemberExportTemplate model
# ---------------------------------------------------------------------------


def test_model_has_required_fields():
    from apps.members.models import MemberExportTemplate

    field_names = {f.name for f in MemberExportTemplate._meta.get_fields()}
    assert field_names >= {
        "name",
        "column_keys",
        "agreement_status_filters",
        "training_groups",
        "created_by",
        "created_at",
        "updated_at",
    }

    name_field = MemberExportTemplate._meta.get_field("name")
    assert name_field.max_length == 128

    col_field = MemberExportTemplate._meta.get_field("column_keys")
    assert col_field.__class__.__name__ == "JSONField"

    ag_field = MemberExportTemplate._meta.get_field("agreement_status_filters")
    assert ag_field.__class__.__name__ == "JSONField"

    tg_field = MemberExportTemplate._meta.get_field("training_groups")
    assert tg_field.many_to_many
    assert tg_field.remote_field.related_name == "export_templates"

    cb_field = MemberExportTemplate._meta.get_field("created_by")
    assert cb_field.many_to_one
    assert cb_field.remote_field.related_name == "member_export_templates_created"


def test_model_clean_rejects_empty_columns():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(name="test", column_keys=[])
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_rejects_invalid_keys():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(name="test", column_keys=["invalid_key"])
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_rejects_duplicate_keys():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(
        name="test", column_keys=["member_full_name", "member_full_name"]
    )
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_rejects_non_string_column_keys():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(name="test", column_keys=[123])
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_rejects_invalid_agreement_states():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["invalid_state"],
    )
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_rejects_duplicate_agreement_states():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed", "signed"],
    )
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_rejects_non_string_agreement_states():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=[1],
    )
    with pytest.raises(ValidationError):
        template.full_clean()


def test_model_clean_accepts_valid_template():
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate(
        name="valid",
        column_keys=["member_full_name", "guardian_email"],
        agreement_status_filters=["signed"],
    )
    template.full_clean()  # must not raise


# ---------------------------------------------------------------------------
# build_template_member_queryset
# ---------------------------------------------------------------------------


def test_build_queryset_empty_filters_unrestricted(member):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name"]
    )
    qs = build_template_member_queryset(template)
    assert qs.count() == Member.objects.count()
    assert qs.filter(pk=member.pk).exists()


def test_build_queryset_filters_by_agreement_state(member):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    _make_agreement(member, state="signed")

    template = MemberExportTemplate.objects.create(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed"],
    )
    assert build_template_member_queryset(template).filter(pk=member.pk).exists()

    template.agreement_status_filters = ["void"]
    template.save()
    assert not build_template_member_queryset(template).filter(pk=member.pk).exists()


def test_build_queryset_filters_agreement_state_or(member):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    _make_agreement(member, state="signed")

    template = MemberExportTemplate.objects.create(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["void", "signed"],
    )
    assert build_template_member_queryset(template).filter(pk=member.pk).exists()


def test_build_queryset_filters_by_training_group(member, training_group_a):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    member.training_group = training_group_a
    member.save(update_fields=["training_group"])

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name"]
    )
    template.training_groups.add(training_group_a)

    assert build_template_member_queryset(template).filter(pk=member.pk).exists()


def test_build_queryset_filters_training_group_or(
    member, training_group_a, training_group_b
):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    member.training_group = training_group_a
    member.save(update_fields=["training_group"])

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name"]
    )
    template.training_groups.add(training_group_a, training_group_b)

    assert build_template_member_queryset(template).filter(pk=member.pk).exists()


def test_build_queryset_filters_and_both(member, training_group_a, training_group_b):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    _make_agreement(member, state="signed")
    member.training_group = training_group_a
    member.save(update_fields=["training_group"])

    template = MemberExportTemplate.objects.create(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed"],
    )
    template.training_groups.add(training_group_a)

    # Matching state AND matching group -> included.
    assert build_template_member_queryset(template).filter(pk=member.pk).exists()

    # Matching state, non-matching group -> excluded.
    template.training_groups.set([training_group_b])
    assert not build_template_member_queryset(template).filter(pk=member.pk).exists()


def test_build_queryset_current_agreement_only(member):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    # A historical (superseded) signed agreement must not satisfy the filter;
    # only the current agreement's state counts.
    _make_agreement(member, state="signed", is_current=False)
    _make_agreement(member, state="generated", is_current=True)

    template = MemberExportTemplate.objects.create(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed"],
    )
    assert not build_template_member_queryset(template).filter(pk=member.pk).exists()


def test_build_queryset_no_duplicate_member_rows(member):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    _make_agreement(member, state="signed", is_current=False)
    _make_agreement(member, state="signed", is_current=True)

    template = MemberExportTemplate.objects.create(
        name="test",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed"],
    )
    pks = list(
        build_template_member_queryset(template).values_list("pk", flat=True)
    )
    assert pks.count(member.pk) == 1
    assert len(pks) == len(set(pks))


def test_build_queryset_group_or_multiple_members(
    guardian, training_group_a, training_group_b
):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    in_a = _member_with_agreement(guardian, group=training_group_a, full_name="InA")
    in_b = _member_with_agreement(guardian, group=training_group_b, full_name="InB")
    in_neither = _member_with_agreement(guardian, group=None, full_name="Neither")

    template = MemberExportTemplate.objects.create(
        name="t", column_keys=["member_full_name"]
    )
    template.training_groups.add(training_group_a, training_group_b)

    pks = set(build_template_member_queryset(template).values_list("pk", flat=True))
    assert in_a.pk in pks
    assert in_b.pk in pks
    assert in_neither.pk not in pks


def test_build_queryset_state_or_only_current_states(guardian):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    signed = _member_with_agreement(guardian, state="signed", full_name="Signed")
    void = _member_with_agreement(guardian, state="void", full_name="Void")
    generated = _member_with_agreement(guardian, state="generated", full_name="Generated")

    template = MemberExportTemplate.objects.create(
        name="t",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed", "void"],
    )
    pks = set(build_template_member_queryset(template).values_list("pk", flat=True))
    assert signed.pk in pks
    assert void.pk in pks
    assert generated.pk not in pks


def test_build_queryset_both_filter_mismatch_excluded(
    guardian, training_group_a, training_group_b
):
    from apps.members.export_templates import build_template_member_queryset
    from apps.members.models import MemberExportTemplate

    match = _member_with_agreement(
        guardian, group=training_group_a, state="signed", full_name="Match"
    )
    wrong_group = _member_with_agreement(
        guardian, group=training_group_b, state="signed", full_name="WrongGroup"
    )
    wrong_state = _member_with_agreement(
        guardian, group=training_group_a, state="void", full_name="WrongState"
    )

    template = MemberExportTemplate.objects.create(
        name="t",
        column_keys=["member_full_name"],
        agreement_status_filters=["signed"],
    )
    template.training_groups.add(training_group_a)

    pks = set(build_template_member_queryset(template).values_list("pk", flat=True))
    assert match.pk in pks
    assert wrong_group.pk not in pks
    assert wrong_state.pk not in pks


# ---------------------------------------------------------------------------
# render_member_export
# ---------------------------------------------------------------------------


def test_render_member_export_csv(member):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name"]
    )
    result = render_member_export(template, "csv")

    assert result.row_count == 1
    assert result.sensitive is False
    assert result.response["Content-Type"] == "text/csv; charset=utf-8"
    assert "attachment" in result.response["Content-Disposition"]
    assert result.response.content[:3] == b"\xef\xbb\xbf"


def test_render_member_export_xlsx(member):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name"]
    )
    result = render_member_export(template, "xlsx")

    assert result.row_count == 1
    assert result.sensitive is False
    assert (
        result.response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in result.response["Content-Disposition"]


def test_render_member_export_sensitive_flag(member):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name", "guardian_email"]
    )
    result = render_member_export(template, "csv")
    assert result.sensitive is True


def test_render_member_export_expected_header_order(member):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["guardian_name", "member_full_name"]
    )
    result = render_member_export(template, "csv")
    body = result.response.content.decode("utf-8")[1:]  # strip BOM
    header = next(csv.reader(io.StringIO(body), delimiter=";"))
    assert header == ["Vecāka vārds, uzvārds", "Biedra vārds, uzvārds"]


def test_render_member_export_guards_raw_values_exactly_once(guardian):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    Member.objects.create(full_name="=EVIL()", guardian=guardian)
    template = MemberExportTemplate.objects.create(
        name="test", column_keys=["member_full_name"]
    )
    result = render_member_export(template, "csv")
    body = result.response.content.decode("utf-8")[1:]
    rows = list(csv.reader(io.StringIO(body), delimiter=";"))
    assert rows[1] == ["'=EVIL()"]


def test_render_member_export_no_n_plus_one(guardian, training_group_a):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    # Setup runs BEFORE the capture window, so its queries are excluded from
    # the count below: three members sharing one guardian and one group.
    for i in range(3):
        Member.objects.create(
            full_name=f"Biedrs {i}", guardian=guardian, training_group=training_group_a
        )
    template = MemberExportTemplate.objects.create(
        name="test",
        column_keys=[
            "member_full_name",
            "guardian_name",
            "agreement_state",
            "training_group_name",
        ],
    )

    # Expected: at most 3 queries — (1) training_groups existence check,
    # (2) select_related member rows, (3) current-agreement Prefetch. A
    # per-member relationship query (N+1) would exceed this bound.
    with CaptureQueriesContext(connection) as ctx:
        result = render_member_export(template, "csv")
    assert result.row_count == 3
    assert len(ctx.captured_queries) <= 3


def test_render_member_export_xlsx_header_order_and_formula_guard(guardian):
    from apps.members.export_templates import render_member_export
    from apps.members.models import MemberExportTemplate

    Member.objects.create(full_name="=EVIL()", guardian=guardian)
    template = MemberExportTemplate.objects.create(
        name="t", column_keys=["guardian_name", "member_full_name"]
    )
    result = render_member_export(template, "xlsx")

    import io as _io

    from openpyxl import load_workbook

    ws = load_workbook(_io.BytesIO(result.response.content)).active
    assert [c.value for c in ws[1]] == [
        "Vecāka vārds, uzvārds",
        "Biedra vārds, uzvārds",
    ]
    # Member full name beginning with '=' is guarded exactly once, not twice.
    assert ws["B2"].value == "'=EVIL()"


# ---------------------------------------------------------------------------
# Forms + ordered-column widget
# ---------------------------------------------------------------------------


def test_ordered_column_keys_widget_provides_available_columns():
    from apps.members.forms import OrderedColumnKeysWidget

    widget = OrderedColumnKeysWidget()
    ctx = widget.get_context("column_keys", ["member_full_name"], {})
    assert "available_keys" in ctx["widget"]
    assert "member_full_name" in ctx["widget"]["selected"]
    assert "guardian_email" in ctx["widget"]["available_keys"]


def test_ordered_column_keys_widget_serializes_selected():
    from apps.members.forms import OrderedColumnKeysWidget

    widget = OrderedColumnKeysWidget()
    ctx = widget.get_context(
        "column_keys", ["member_full_name", "guardian_name"], {}
    )
    assert ctx["widget"]["selected"] == ["member_full_name", "guardian_name"]
    rendered = widget.render("column_keys", ["member_full_name"])
    assert 'type="hidden"' in rendered


def test_ordered_column_keys_widget_has_move_and_remove_hooks():
    from apps.members.forms import OrderedColumnKeysWidget

    widget = OrderedColumnKeysWidget()
    rendered = widget.render("column_keys", ["member_full_name"])
    # add/remove/move-up/move-down + hidden JSON hooks all present.
    assert "fk-column-add" in rendered
    assert "fk-column-remove" in rendered
    assert "fk-column-move-up" in rendered
    assert "fk-column-move-down" in rendered
    assert 'name="column_keys"' in rendered


def test_member_export_template_admin_form_validates_column_keys():
    from apps.members.forms import MemberExportTemplateAdminForm

    form = MemberExportTemplateAdminForm(
        data={"name": "Test", "column_keys": '["invalid_key"]'}
    )
    assert not form.is_valid()
    assert "column_keys" in form.errors


def test_member_export_template_admin_form_validates_agreement_states():
    from apps.members.forms import MemberExportTemplateAdminForm

    form = MemberExportTemplateAdminForm(
        data={
            "name": "Test",
            "column_keys": '["member_full_name"]',
            "agreement_status_filters": ["invalid_state"],
        }
    )
    assert not form.is_valid()
    assert "agreement_status_filters" in form.errors


def test_member_export_template_admin_form_rejects_empty_columns():
    from apps.members.forms import MemberExportTemplateAdminForm

    form = MemberExportTemplateAdminForm(
        data={"name": "Test", "column_keys": "[]"}
    )
    assert not form.is_valid()
    assert "column_keys" in form.errors


def test_member_export_template_admin_form_rejects_duplicate_column_keys():
    from apps.members.forms import MemberExportTemplateAdminForm

    form = MemberExportTemplateAdminForm(
        data={
            "name": "Test",
            "column_keys": '["member_full_name", "member_full_name"]',
        }
    )
    assert not form.is_valid()
    assert "column_keys" in form.errors


def test_member_export_template_admin_form_rejects_duplicate_agreement_states():
    from apps.members.forms import MemberExportTemplateAdminForm

    form = MemberExportTemplateAdminForm(
        data={
            "name": "Test",
            "column_keys": '["member_full_name"]',
            "agreement_status_filters": ["signed", "signed"],
        }
    )
    assert not form.is_valid()
    assert "agreement_status_filters" in form.errors


def test_member_export_run_form_xlsx_initial():
    from apps.members.forms import MemberExportRunForm

    form = MemberExportRunForm()
    assert form.initial["fmt"] == "xlsx"
