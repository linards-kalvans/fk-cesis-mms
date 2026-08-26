"""csv_response — UTF-8 BOM + semicolon CSV with value formatting + injection guard.

P17 adds ``prepare_export_cell`` (shared cell formatting + guard) and
``xlsx_response`` (in-memory XLSX writer).
"""

import csv
import datetime
import io

from apps.core.export import csv_response


def _parse(resp):
    body = resp.content.decode("utf-8")
    assert body.startswith("﻿")  # UTF-8 BOM
    reader = csv.reader(io.StringIO(body[1:]), delimiter=";")
    return list(reader)


def test_header_and_delimiter_and_bom():
    resp = csv_response(filename="x.csv", columns=["A", "B"], rows=[["1", "2"]])
    assert resp["Content-Type"].startswith("text/csv")
    assert 'attachment; filename="x.csv"' in resp["Content-Disposition"]
    rows = _parse(resp)
    assert rows[0] == ["A", "B"]
    assert rows[1] == ["1", "2"]


def test_value_formatting():
    resp = csv_response(
        filename="x.csv",
        columns=["a", "b", "c", "d"],
        rows=[[None, True, False, datetime.date(2026, 9, 20)]],
    )
    assert _parse(resp)[1] == ["", "jā", "nē", "2026-09-20"]


def test_datetime_formatting():
    resp = csv_response(
        filename="x.csv",
        columns=["ts"],
        rows=[[datetime.datetime(2026, 9, 20, 14, 30, 5)]],
    )
    assert _parse(resp)[1] == ["2026-09-20 14:30"]


def test_formula_injection_guard():
    resp = csv_response(filename="x.csv", columns=["a"], rows=[["=SUM(A1:A2)"], ["+1"], ["@x"], ["-3"], ["safe"]])
    data = _parse(resp)
    assert data[1] == ["'=SUM(A1:A2)"]
    assert data[2] == ["'+1"]
    assert data[3] == ["'@x"]
    assert data[4] == ["'-3"]
    assert data[5] == ["safe"]


# ---------------------------------------------------------------------------
# P17 — shared cell formatting + XLSX writer
# ---------------------------------------------------------------------------


def test_prepare_export_cell_formats_none():
    from apps.core.export import prepare_export_cell

    assert prepare_export_cell(None) == ""


def test_prepare_export_cell_formats_bool():
    from apps.core.export import prepare_export_cell

    assert prepare_export_cell(True) == "jā"
    assert prepare_export_cell(False) == "nē"


def test_prepare_export_cell_formats_date():
    from apps.core.export import prepare_export_cell

    assert prepare_export_cell(datetime.date(2026, 8, 26)) == "2026-08-26"


def test_prepare_export_cell_guard_formula_all_sentinels():
    from apps.core.export import prepare_export_cell

    assert prepare_export_cell("=SUM(A1)") == "'=SUM(A1)"
    assert prepare_export_cell("+1") == "'+1"
    assert prepare_export_cell("-1") == "'-1"
    assert prepare_export_cell("@test") == "'@test"
    assert prepare_export_cell("\ttab") == "'\ttab"
    assert prepare_export_cell("\rCR") == "'\rCR"
    assert prepare_export_cell("normal") == "normal"


def test_xlsx_response_mime_and_attachment():
    from apps.core.export import xlsx_response

    resp = xlsx_response(filename="x.xlsx", columns=["A"], rows=[["1"]])
    assert (
        resp["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'attachment; filename="x.xlsx"' in resp["Content-Disposition"]
    # XLSX is a ZIP archive — begins with the PK local-file-header magic.
    assert resp.content[:2] == b"PK"


def test_xlsx_response_readable_workbook_with_bold_headers():
    from apps.core.export import xlsx_response

    import io as _io

    from openpyxl import load_workbook

    resp = xlsx_response(filename="x.xlsx", columns=["A", "B"], rows=[["1", "2"]])
    wb = load_workbook(_io.BytesIO(resp.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["A", "B"]
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    assert [c.value for c in ws[2]] == ["1", "2"]


def test_xlsx_response_bounds_column_widths():
    from apps.core.export import xlsx_response

    import io as _io

    from openpyxl import load_workbook

    short = xlsx_response(filename="s.xlsx", columns=["A"], rows=[["x"]])
    ws = load_workbook(_io.BytesIO(short.content)).active
    assert ws.column_dimensions["A"].width >= 10

    long = xlsx_response(filename="l.xlsx", columns=["A" * 50], rows=[["x"]])
    ws = load_workbook(_io.BytesIO(long.content)).active
    assert ws.column_dimensions["A"].width <= 40


def test_xlsx_response_applies_formula_guard():
    from apps.core.export import xlsx_response

    import io as _io

    from openpyxl import load_workbook

    resp = xlsx_response(
        filename="g.xlsx", columns=["a"], rows=[["=SUM(A1)"], ["safe"]]
    )
    ws = load_workbook(_io.BytesIO(resp.content)).active
    assert ws["A2"].value == "'=SUM(A1)"
    assert ws["A3"].value == "safe"
